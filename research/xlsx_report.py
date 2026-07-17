"""
Excel 报告生成
==============

用 openpyxl 生成美观的 .xlsx 报告:
- 单因子 Sheet: 绩效指标表 + 月度收益 + 净值曲线 + 回撤 + IC + 分层
- 多因子对比 Sheet: 指标对比表 + 排名 + 净值对比图
- 条件格式: 收益正绿负红、数据条、色阶

依赖: openpyxl + matplotlib
"""
from __future__ import annotations

from pathlib import Path
import tempfile

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as XlImage
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backtest.engine import BacktestResult
from backtest.metrics import METRIC_LABELS, calc_all_metrics

# 中文字体
for _font in ["Microsoft YaHei", "SimHei", "WenQuanYi Micro Hei", "Arial Unicode MS"]:
    try:
        fm.findfont(_font, fallback_to_default=False)
        plt.rcParams["font.sans-serif"] = [_font, "DejaVu Sans"]
        plt.rcParams["axes.unicode_minus"] = False
        break
    except Exception:
        continue
else:
    plt.rcParams["axes.unicode_minus"] = False

# ===========================================================================
# 样式定义
# ===========================================================================
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(name="微软雅黑", bold=True, color="1F4E79", size=16)
_SUBTITLE_FONT = Font(name="微软雅黑", bold=True, color="1F4E79", size=12)
_LABEL_FONT = Font(name="微软雅黑", size=10, color="333333")
_VALUE_FONT = Font(name="Consolas", size=10, color="333333")
_POS_FONT = Font(name="Consolas", size=10, color="008000", bold=True)
_NEG_FONT = Font(name="Consolas", size=10, color="CC0000", bold=True)
_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
_ALT_FILL = PatternFill("solid", fgColor="F2F6FC")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", indent=1)
_RIGHT = Alignment(horizontal="right", vertical="center")


def _apply_header(ws, row, cols):
    """给表头行加样式。"""
    for c in cols:
        cell = ws[f"{c}{row}"]
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def _write_value(ws, row, col, value, pct=False, color_pos=False):
    """写入数值，自动格式化。"""
    cell = ws[f"{col}{row}"]
    if isinstance(value, float):
        if pct:
            cell.value = value
            cell.number_format = "0.00%"
        else:
            cell.value = round(value, 4)
            cell.number_format = "0.0000"
    else:
        cell.value = value
    cell.font = _VALUE_FONT
    cell.alignment = _RIGHT
    cell.border = _BORDER
    if color_pos and isinstance(value, float):
        cell.font = _POS_FONT if value >= 0 else _NEG_FONT
    if row % 2 == 0:
        cell.fill = _ALT_FILL


# ===========================================================================
# 图表生成（matplotlib → PNG → 插入 Excel）
# ===========================================================================
def _save_equity_png(result: BacktestResult, path: Path):
    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.plot(result.equity_curve.index, result.equity_curve.values, linewidth=1.5, color="#1F4E79")
    ax.fill_between(result.equity_curve.index, 1, result.equity_curve.values, alpha=0.08, color="#1F4E79")
    ax.set_title("净值曲线 Equity Curve", fontsize=12)
    ax.set_ylabel("NAV")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_drawdown_png(result: BacktestResult, path: Path):
    cum = (1 + result.daily_returns).cumprod()
    dd = (cum - cum.cummax()) / cum.cummax()
    fig, ax = plt.subplots(figsize=(10, 2.8))
    ax.fill_between(dd.index, dd.values, 0, color="#CC0000", alpha=0.35)
    ax.set_title("回撤 Drawdown", fontsize=12)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_ic_png(ic_series: pd.Series, path: Path):
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 5), height_ratios=[2, 1])
    ax1.bar(ic_series.index, ic_series.values, width=1, alpha=0.6, color="steelblue")
    ax1.axhline(y=0, color="black", linewidth=0.5)
    ax1.set_ylabel("Daily IC")
    ax1.grid(True, alpha=0.3)
    cum = ic_series.cumsum()
    ax2.fill_between(cum.index, 0, cum.values, alpha=0.2, color="darkblue")
    ax2.plot(cum.index, cum.values, linewidth=1, color="darkblue")
    ax2.set_ylabel("Cumulative IC")
    ax2.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_layers_png(layer_nav: pd.DataFrame, path: Path):
    fig, ax = plt.subplots(figsize=(10, 4))
    for col in layer_nav.columns:
        ax.plot(layer_nav.index, layer_nav[col].values, label=col, linewidth=1)
    ax.set_title("分层回测 Quantile NAV", fontsize=12)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_monthly_png(daily_returns: pd.Series, path: Path):
    monthly = (1 + daily_returns).resample("ME").apply(lambda x: (1 + x).prod() - 1)
    if monthly.empty:
        return
    pivot = monthly.to_frame("ret")
    pivot["y"] = pivot.index.year
    pivot["m"] = pivot.index.month
    table = pivot.pivot_table(index="y", columns="m", values="ret").reindex(columns=range(1, 13))
    fig, ax = plt.subplots(figsize=(10, 3.5))
    im = ax.imshow(table.values, cmap="RdYlGn", aspect="auto", vmin=-0.08, vmax=0.08)
    ax.set_xticks(range(12))
    ax.set_xticklabels([f"{m}月" for m in range(1, 13)], fontsize=8)
    ax.set_yticks(range(len(table.index)))
    ax.set_yticklabels(table.index, fontsize=8)
    for i in range(table.shape[0]):
        for j in range(table.shape[1]):
            v = table.iloc[i, j]
            if pd.notna(v):
                ax.text(j, i, f"{v:.1%}", ha="center", va="center", fontsize=6)
    plt.colorbar(im, ax=ax, fraction=0.02)
    ax.set_title("月度收益 Monthly Returns", fontsize=12)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_comparison_equity_png(results: dict[str, BacktestResult], path: Path):
    fig, ax = plt.subplots(figsize=(10, 5))
    colors = plt.cm.tab10(np.linspace(0, 1, len(results)))
    for (name, res), c in zip(results.items(), colors):
        ax.plot(res.equity_curve.index, res.equity_curve.values, label=name, linewidth=1.2, color=c)
    ax.set_title("净值曲线对比 Equity Comparison", fontsize=12)
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()


# ===========================================================================
# 单因子 Sheet
# ===========================================================================
def _write_single_sheet(wb: Workbook, factor_name: str, result: BacktestResult, factor_summary: dict | None):
    ws = wb.create_sheet(factor_name[:31])  # Excel sheet 名最长 31

    # ---- 标题 ----
    ws.merge_cells("A1:H1")
    ws["A1"] = f"YuriQuant 回测报告 — {factor_name}"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # ---- 绩效指标表 ----
    metrics = result.metrics()
    row = 3
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "绩效指标 Performance Metrics"
    ws[f"A{row}"].font = _SUBTITLE_FONT
    row += 1

    # 两列布局: 指标名 | 值 | 指标名 | 值
    items = list(metrics.items())
    half = (len(items) + 1) // 2
    for i in range(half):
        r = row + i
        # 左列
        k1 = items[i][0]
        ws[f"A{r}"] = METRIC_LABELS.get(k1, k1)
        ws[f"A{r}"].font = _LABEL_FONT
        ws[f"A{r}"].alignment = _LEFT
        ws[f"A{r}"].border = _BORDER
        v1 = items[i][1]
        ws[f"B{r}"] = v1
        _format_metric_cell(ws[f"B{r}"], k1, v1)
        if r % 2 == 0:
            ws[f"A{r}"].fill = _ALT_FILL
        # 右列
        idx2 = i + half
        if idx2 < len(items):
            k2 = items[idx2][0]
            ws[f"E{r}"] = METRIC_LABELS.get(k2, k2)
            ws[f"E{r}"].font = _LABEL_FONT
            ws[f"E{r}"].alignment = _LEFT
            ws[f"E{r}"].border = _BORDER
            v2 = items[idx2][1]
            ws[f"F{r}"] = v2
            _format_metric_cell(ws[f"F{r}"], k2, v2)
            if r % 2 == 0:
                ws[f"E{r}"].fill = _ALT_FILL
    row += half + 1

    # ---- 因子分析 ----
    if factor_summary:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = "因子分析 Factor Analysis"
        ws[f"A{row}"].font = _SUBTITLE_FONT
        row += 1
        for key in ["ic_mean", "ic_std", "ic_win_rate", "ir"]:
            v = factor_summary.get(key)
            if v is not None:
                ws[f"A{row}"] = METRIC_LABELS.get(key, key)
                ws[f"A{row}"].font = _LABEL_FONT
                ws[f"A{row}"].alignment = _LEFT
                ws[f"A{row}"].border = _BORDER
                ws[f"B{row}"] = v
                _format_metric_cell(ws[f"B{row}"], key, v)
                row += 1
        # IC 衰减
        if "ic_decay" in factor_summary:
            ws[f"A{row}"] = "IC衰减 IC Decay"
            ws[f"A{row}"].font = _LABEL_FONT
            ws[f"A{row}"].alignment = _LEFT
            ws[f"A{row}"].border = _BORDER
            decay_str = ", ".join(f"lag{k}={v:.4f}" for k, v in factor_summary["ic_decay"].items())
            ws.merge_cells(f"B{row}:H{row}")
            ws[f"B{row}"] = decay_str
            ws[f"B{row}"].font = _VALUE_FONT
            ws[f"B{row}"].alignment = _LEFT
            row += 1
        # 分层收益
        if "layer_returns" in factor_summary:
            ws[f"A{row}"] = "分层累计收益 Layer Returns"
            ws[f"A{row}"].font = _LABEL_FONT
            ws[f"A{row}"].alignment = _LEFT
            ws[f"A{row}"].border = _BORDER
            layer_str = ", ".join(f"{q}={v:.4f}" for q, v in factor_summary["layer_returns"].items())
            ws.merge_cells(f"B{row}:H{row}")
            ws[f"B{row}"] = layer_str
            ws[f"B{row}"].font = _VALUE_FONT
            ws[f"B{row}"].alignment = _LEFT
            row += 1
    row += 2

    # ---- 嵌入图表 ----
    tmp = Path(tempfile.gettempdir()) / "yuriquant"
    tmp.mkdir(exist_ok=True)

    # 净值
    p = tmp / f"equity_{factor_name}.png"
    _save_equity_png(result, p)
    img = XlImage(str(p))
    img.width, img.height = 680, 300
    ws.add_image(img, f"A{row}")
    row += 16

    # 回撤
    p = tmp / f"dd_{factor_name}.png"
    _save_drawdown_png(result, p)
    img = XlImage(str(p))
    img.width, img.height = 680, 190
    ws.add_image(img, f"A{row}")
    row += 11

    # 月度热力图
    p = tmp / f"monthly_{factor_name}.png"
    _save_monthly_png(result.daily_returns, p)
    img = XlImage(str(p))
    img.width, img.height = 680, 240
    ws.add_image(img, f"A{row}")
    row += 13

    # IC
    if factor_summary and "ic_series" in factor_summary:
        p = tmp / f"ic_{factor_name}.png"
        _save_ic_png(factor_summary["ic_series"], p)
        img = XlImage(str(p))
        img.width, img.height = 680, 340
        ws.add_image(img, f"A{row}")
        row += 18

    # 分层
    if factor_summary and "layer_nav" in factor_summary:
        p = tmp / f"layer_{factor_name}.png"
        _save_layers_png(factor_summary["layer_nav"], p)
        img = XlImage(str(p))
        img.width, img.height = 680, 270
        ws.add_image(img, f"A{row}")

    # ---- 月度收益数据表 ----
    row += 18
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = "月度收益明细 Monthly Returns Detail"
    ws[f"A{row}"].font = _SUBTITLE_FONT
    row += 1
    monthly = (1 + result.daily_returns).resample("ME").apply(lambda x: (1 + x).prod() - 1)
    if not monthly.empty:
        # 表头
        ws[f"A{row}"] = "年份 Year"
        _apply_header(ws, row, ["A"])
        for m in range(1, 13):
            col = get_column_letter(m + 1)
            ws[f"{col}{row}"] = f"{m}月"
            _apply_header(ws, row, [col])
        ws[f"{get_column_letter(14)}{row}"] = "全年"
        _apply_header(ws, row, [get_column_letter(14)])
        row += 1
        # 数据
        for year in sorted(set(monthly.index.year)):
            ws[f"A{row}"] = year
            ws[f"A{row}"].font = _LABEL_FONT
            ws[f"A{row}"].border = _BORDER
            ws[f"A{row}"].alignment = _CENTER
            for m in range(1, 13):
                col = get_column_letter(m + 1)
                mask = (monthly.index.year == year) & (monthly.index.month == m)
                if mask.any():
                    val = monthly[mask].iloc[0]
                    ws[f"{col}{row}"] = val
                    ws[f"{col}{row}"].number_format = "0.00%"
                    ws[f"{col}{row}"].font = _POS_FONT if val >= 0 else _NEG_FONT
                ws[f"{col}{row}"].border = _BORDER
                ws[f"{col}{row}"].alignment = _RIGHT
            # 全年
            yr_ret = (1 + monthly[monthly.index.year == year]).prod() - 1
            ws[f"{get_column_letter(14)}{row}"] = yr_ret
            ws[f"{get_column_letter(14)}{row}"].number_format = "0.00%"
            ws[f"{get_column_letter(14)}{row}"].font = _POS_FONT if yr_ret >= 0 else _NEG_FONT
            ws[f"{get_column_letter(14)}{row}"].border = _BORDER
            ws[f"{get_column_letter(14)}{row}"].alignment = _RIGHT
            row += 1
        # 条件格式: 月度收益色阶
        cs_rule = ColorScaleRule(
            start_type="num", start_value=-0.08, start_color="CC0000",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=0.08, end_color="008000",
        )
        ws.conditional_formatting.add(f"B{row - len(set(monthly.index.year))}:M{row - 1}", cs_rule)

    # 列宽
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 16
    for c in "GHIJ":
        ws.column_dimensions[c].width = 4
    for m in range(1, 15):
        ws.column_dimensions[get_column_letter(m)].width = 10 if m > 1 else 28


def _format_metric_cell(cell, key: str, value):
    """格式化指标值单元格。"""
    if isinstance(value, float):
        if key in ("win_rate", "ic_win_rate"):
            cell.number_format = "0.00%"
        elif abs(value) < 100:
            cell.number_format = "0.0000"
        else:
            cell.number_format = "0"
    cell.font = _VALUE_FONT
    cell.alignment = _RIGHT
    cell.border = _BORDER
    # 正负色
    if key in ("annual_return", "total_return", "sharpe", "sortino", "calmar",
               "excess_return", "information_ratio", "ir", "ic_mean",
               "avg_daily_return", "profit_loss_ratio"):
        if isinstance(value, float):
            cell.font = _POS_FONT if value >= 0 else _NEG_FONT


# ===========================================================================
# 多因子对比 Sheet
# ===========================================================================
def _write_comparison_sheet(wb: Workbook, results: dict[str, BacktestResult], factor_summaries: dict[str, dict] | None):
    ws = wb.create_sheet("对比 Comparison")

    # 标题
    ws.merge_cells("A1:L1")
    ws["A1"] = "YuriQuant 多因子对比报告 Multi-Factor Comparison"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # ---- 指标对比表 ----
    row = 3
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "指标对比 Metrics Comparison"
    ws[f"A{row}"].font = _SUBTITLE_FONT
    row += 1

    # 收集指标
    key_metrics = [
        "annual_return", "annual_volatility", "sharpe", "sortino",
        "max_drawdown", "calmar", "win_rate", "profit_loss_ratio",
        "avg_turnover", "ic_mean", "ir",
    ]
    headers = ["指标 Metric"] + list(results.keys())
    for c, h in enumerate(headers, 1):
        ws[f"{get_column_letter(c)}{row}"] = h
    _apply_header(ws, row, [get_column_letter(c) for c in range(1, len(headers) + 1)])
    row += 1

    for k in key_metrics:
        ws[f"A{row}"] = METRIC_LABELS.get(k, k)
        ws[f"A{row}"].font = _LABEL_FONT
        ws[f"A{row}"].alignment = _LEFT
        ws[f"A{row}"].border = _BORDER
        for c, (name, res) in enumerate(results.items(), 2):
            m = res.metrics()
            if k in m:
                v = m[k]
            elif factor_summaries and name in factor_summaries:
                v = factor_summaries[name].get(k, None)
            else:
                v = None
            cell = ws[f"{get_column_letter(c)}{row}"]
            if v is not None:
                cell.value = v
                _format_metric_cell(cell, k, v)
            else:
                cell.value = "-"
                cell.font = _VALUE_FONT
                cell.alignment = _RIGHT
                cell.border = _BORDER
            if row % 2 == 0:
                ws[f"A{row}"].fill = _ALT_FILL
        row += 1

    # 条件格式: 夏普和年化收益用数据条
    n_cols = len(results)
    # 找到 sharpe 和 annual_return 的行号
    for k_idx, k in enumerate(key_metrics):
        if k in ("sharpe", "annual_return", "ir"):
            r = row - len(key_metrics) + k_idx
            data_range = f"B{r}:{get_column_letter(n_cols + 1)}{r}"
            ws.conditional_formatting.add(data_range, DataBarRule(
                start_type="min", end_type="max", color="5B9BD5"
            ))

    row += 2

    # ---- 排名表 ----
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "排名 Ranking (by Sharpe)"
    ws[f"A{row}"].font = _SUBTITLE_FONT
    row += 1

    rank_headers = ["排名 Rank", "因子 Factor", "年化收益 Annual Return", "夏普 Sharpe", "最大回撤 Max DD", "IC均值 IC Mean", "IR"]
    for c, h in enumerate(rank_headers, 1):
        ws[f"{get_column_letter(c)}{row}"] = h
    _apply_header(ws, row, [get_column_letter(c) for c in range(1, len(rank_headers) + 1)])
    row += 1

    # 按夏普降序
    ranked = sorted(results.items(), key=lambda x: x[1].metrics()["sharpe"], reverse=True)
    for rank, (name, res) in enumerate(ranked, 1):
        m = res.metrics()
        ic_mean = factor_summaries.get(name, {}).get("ic_mean", None) if factor_summaries else None
        ir_val = factor_summaries.get(name, {}).get("ir", None) if factor_summaries else None
        ws[f"A{row}"] = rank
        ws[f"B{row}"] = name
        ws[f"C{row}"] = m["annual_return"]; ws[f"C{row}"].number_format = "0.00%"; ws[f"C{row}"].font = _POS_FONT if m["annual_return"] >= 0 else _NEG_FONT
        ws[f"D{row}"] = m["sharpe"]; ws[f"D{row}"].number_format = "0.0000"; ws[f"D{row}"].font = _POS_FONT if m["sharpe"] >= 0 else _NEG_FONT
        ws[f"E{row}"] = m["max_drawdown"]; ws[f"E{row}"].number_format = "0.00%"
        ws[f"F{row}"] = ic_mean if ic_mean is not None else "-"; ws[f"F{row}"].number_format = "0.0000" if ic_mean is not None else "General"
        ws[f"G{row}"] = ir_val if ir_val is not None else "-"; ws[f"G{row}"].number_format = "0.0000" if ir_val is not None else "General"
        for c in range(1, len(rank_headers) + 1):
            ws[f"{get_column_letter(c)}{row}"].border = _BORDER
            ws[f"{get_column_letter(c)}{row}"].alignment = _CENTER if c <= 2 else _RIGHT
            if c > 2 and c not in (3, 4):
                ws[f"{get_column_letter(c)}{row}"].font = _VALUE_FONT
            if row % 2 == 0:
                ws[f"{get_column_letter(c)}{row}"].fill = _ALT_FILL
        row += 1

    row += 2

    # ---- 净值对比图 ----
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "净值曲线对比 Equity Curve Comparison"
    ws[f"A{row}"].font = _SUBTITLE_FONT
    row += 1

    tmp = Path(tempfile.gettempdir()) / "yuriquant"
    tmp.mkdir(exist_ok=True)
    p = tmp / "comparison_equity.png"
    _save_comparison_equity_png(results, p)
    img = XlImage(str(p))
    img.width, img.height = 700, 360
    ws.add_image(img, f"A{row}")

    # 列宽
    ws.column_dimensions["A"].width = 35
    for c in range(2, n_cols + 2):
        ws.column_dimensions[get_column_letter(c)].width = 18


# ===========================================================================
# 主入口
# ===========================================================================
def generate_excel_report(
    results: dict[str, BacktestResult],
    factor_summaries: dict[str, dict] | None = None,
    output_path: str | Path = "reports/yuriquant_report.xlsx",
) -> Path:
    """生成美观的 Excel 报告。

    Args:
        results: {factor_name: BacktestResult}
        factor_summaries: {factor_name: factor_summary dict}
        output_path: 输出路径
    Returns:
        Path to the xlsx file.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    # 单因子 sheets
    for name, result in results.items():
        fs = factor_summaries.get(name, {}) if factor_summaries else None
        _write_single_sheet(wb, name, result, fs)

    # 多因子对比 sheet
    if len(results) > 1:
        _write_comparison_sheet(wb, results, factor_summaries)

    wb.save(out)
    return out
